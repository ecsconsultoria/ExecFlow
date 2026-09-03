"""tests/test_financial_exports_etapa12e.py — Exportação financeira (Etapa 12E-A3).

Cobre, para cada tela: tela continua 200 com botões; PDF/XLSX 200 com
mimetype correto; dados e filtros corretos; RBAC/autenticação; isolamento
por empresa; UTF-8/BRL; ausência de escrita no banco.
"""
import uuid
from datetime import date, datetime, timedelta
from io import BytesIO

import pytest
from openpyxl import load_workbook

from app import create_app
from app.extensions import db
from app.models.client import Client
from app.models.company import Company
from app.models.financial import FinancialRecord
from app.models.financial_catalog import CostCenter, FinancialCategory
from app.models.order import Order, OrderPayment
from app.models.supplier import Supplier
from app.models.user import User

ADMIN_EMAIL = "admin@executivecarsp.com"
ADMIN_PWD = "admin123"


@pytest.fixture(scope="session")
def testing_app():
    return create_app("testing")


@pytest.fixture(autouse=True)
def clean_tables(testing_app):
    with testing_app.app_context():
        for model in (FinancialRecord, OrderPayment, Order,
                      CostCenter, FinancialCategory, Client, Supplier):
            model.query.delete()
        db.session.commit()
    yield


def _login(app, email=ADMIN_EMAIL, password=ADMIN_PWD):
    c = app.test_client()
    r = c.post("/auth/login", data={"email": email, "password": password},
               follow_redirects=False)
    assert r.status_code in (200, 302)
    return c


def _cid_of(app, email=ADMIN_EMAIL):
    with app.app_context():
        return User.query.filter_by(email=email).first().company_id


def _seed_catalog(app, cid):
    with app.app_context():
        pessoal = FinancialCategory(company_id=cid, name="Pessoal",
                                    type="expense", active=True)
        db.session.add(pessoal)
        db.session.flush()
        prolabore = FinancialCategory(company_id=cid, name="Pró-Labore",
                                      type="expense", active=True,
                                      parent_id=pessoal.id)
        impostos = FinancialCategory(company_id=cid, name="Impostos e Tributos",
                                     type="expense", active=True)
        db.session.add_all([prolabore, impostos])
        db.session.flush()
        das = FinancialCategory(company_id=cid, name="DAS / Simples Nacional",
                                type="expense", active=True, parent_id=impostos.id)
        cc = CostCenter(company_id=cid, name="Administrativo", active=True)
        db.session.add_all([das, cc])
        db.session.commit()
        return {"prolabore": prolabore.id, "impostos": impostos.id,
                "das": das.id, "cc": cc.id}


def _seed_expense(app, cid, cat_id, cc_id, *, desc, amount, emission,
                  due, status="pendente", paid_date=None, supplier_id=None):
    with app.app_context():
        fr = FinancialRecord(company_id=cid, type="expense", category="outro",
                             description=desc, amount=amount, status=status,
                             emission_date=emission, due_date=due,
                             paid_date=paid_date,
                             financial_category_id=cat_id, cost_center_id=cc_id,
                             supplier_id=supplier_id,
                             reference=f"expense:{uuid.uuid4().hex[:6]}",
                             notes=f"Observação — {desc}")
        db.session.add(fr)
        db.session.commit()
        return fr.id


def _seed_supplier(app, cid, name):
    with app.app_context():
        s = Supplier(company_id=cid, name=name)
        db.session.add(s)
        db.session.commit()
        return s.id


def _xlsx_rows(resp):
    wb = load_workbook(BytesIO(resp.data))
    ws = wb.active
    return ws, [[c.value for c in row] for row in ws.iter_rows()]


def _row_of(ws_rows, first_cell_value):
    for row in ws_rows:
        if row and row[0] == first_cell_value:
            return row
    return None


# ════════════════════════ FLUXO DE CAIXA ════════════════════════

def test_cash_flow_screen_and_exports(testing_app):
    cid = _cid_of(testing_app)
    cats = _seed_catalog(testing_app, cid)
    _seed_expense(testing_app, cid, cats["prolabore"], cats["cc"],
                  desc="Pró-Labore — competência 07/2026", amount=10000.0,
                  emission=date(2026, 7, 31), due=date(2026, 8, 10),
                  status="pago", paid_date=date(2026, 8, 10))
    _seed_expense(testing_app, cid, cats["das"], cats["cc"],
                  desc="DAS / Simples Nacional — competência 07/2026",
                  amount=1330.21, emission=date(2026, 7, 31), due=date(2026, 8, 31))

    c = _login(testing_app)
    r = c.get("/financial/cash-flow")
    assert r.status_code == 200
    html = r.get_data(as_text=True)
    assert "/financial/cash-flow/export/pdf" in html
    assert "/financial/cash-flow/export/xlsx" in html

    rp = c.get("/financial/cash-flow/export/pdf?period=custom&date_from=2026-08-01&date_to=2026-08-31")
    assert rp.status_code == 200 and rp.mimetype == "application/pdf"
    assert b"Fluxo de Caixa" in rp.data
    assert b"SALDO INICIAL" in rp.data
    assert b"10.000" in rp.data  # pró-labore realizado (saída paga)

    rx = c.get("/financial/cash-flow/export/xlsx?period=custom&date_from=2026-08-01&date_to=2026-08-31")
    ws, rows = _xlsx_rows(rx)
    assert rows[0][0] == "Fluxo de Caixa"
    all_rows = [r for r in rows if r and len(r) > 1 and r[1] is not None]
    realized = [r for r in all_rows if r[6] == "Realizado"]
    forecast = [r for r in all_rows if r[6] == "Previsto"]
    assert any("Pró-Labore" in str(r[1]) and r[5] == 10000.0 for r in realized)
    assert any("DAS / Simples Nacional" in str(r[1]) and r[5] == pytest.approx(1330.21)
               for r in forecast)
    # coluna monetária numérica
    money = [r for r in all_rows if r[5] is not None][0]
    for row in ws.iter_rows():
        if row[5].value not in (None,) and isinstance(row[5].value, (int, float)):
            assert "R$" in (row[5].number_format or "")
            break


# ════════════════════════ CONTAS A RECEBER ════════════════════════

def test_receivables_exports(testing_app):
    cid = _cid_of(testing_app)
    with testing_app.app_context():
        client = Client(company_id=cid, name=f"Cliente AR {uuid.uuid4().hex[:4]}")
        db.session.add(client)
        db.session.flush()
        o = Order(company_id=cid, client_id=client.id,
                  number=f"SO-ARX-{uuid.uuid4().hex[:6]}", status="aberto",
                  client_name=client.name, contact_name="", email="", celular="",
                  language="pt", billing_type="recibo", total_amount=2000.0,
                  payment_method="PIX", emission_date=date.today(),
                  created_by=1)
        db.session.add(o)
        db.session.flush()
        pmt = OrderPayment(order_id=o.id, installment_no=1, amount=2000.0,
                           due_date=date(2026, 8, 5), paid_amount=500.0)
        db.session.add(pmt)
        db.session.commit()
        client_name = client.name

    c = _login(testing_app)
    r = c.get("/financial/receivables")
    assert r.status_code == 200
    assert "/financial/receivables/export/pdf" in r.get_data(as_text=True)

    rp = c.get("/financial/receivables/export/pdf?period=all")
    assert rp.status_code == 200 and rp.mimetype == "application/pdf"
    assert b"Contas a Receber" in rp.data
    assert client_name.encode() in rp.data

    rx = c.get("/financial/receivables/export/xlsx?period=all")
    ws, rows = _xlsx_rows(rx)
    data_rows = [r for r in rows if r and r[0] == client_name]
    assert data_rows
    row = data_rows[0]
    assert row[3] == 2000.0   # valor original (numérico)
    assert row[4] == 500.0    # recebido
    assert row[5] == 1500.0   # saldo
    assert row[6] == "Vencido"


# ════════════════════════ CONTAS A PAGAR ════════════════════════

def test_payables_exports_and_supplier_filter(testing_app):
    cid = _cid_of(testing_app)
    cats = _seed_catalog(testing_app, cid)
    sup_a = _seed_supplier(testing_app, cid, "Fornecedor Alfa")
    sup_b = _seed_supplier(testing_app, cid, "Fornecedor Beta")
    _seed_expense(testing_app, cid, cats["das"], cats["cc"],
                  desc="DAS / Simples Nacional — competência 07/2026",
                  amount=1330.21, emission=date(2026, 7, 31), due=date.today() + timedelta(days=10),
                  supplier_id=sup_a)
    _seed_expense(testing_app, cid, cats["prolabore"], cats["cc"],
                  desc="Outra despesa", amount=50.0,
                  emission=date(2026, 8, 1), due=date(2026, 9, 1),
                  supplier_id=sup_b)

    c = _login(testing_app)
    r = c.get("/financial/payables")
    assert r.status_code == 200
    assert "/financial/payables/export/pdf" in r.get_data(as_text=True)

    rp = c.get("/financial/payables/export/pdf?period=all")
    assert rp.status_code == 200 and rp.mimetype == "application/pdf"
    assert b"Contas a Pagar" in rp.data
    assert b"DAS" in rp.data and b"Simples" in rp.data and b"Nacional" in rp.data

    rx = c.get(f"/financial/payables/export/xlsx?period=all&supplier={sup_a}")
    ws, rows = _xlsx_rows(rx)
    descs = [str(r[1]) for r in rows if r and r[1] is not None]
    assert any("DAS / Simples Nacional" in d for d in descs)
    assert not any("Outra despesa" in d for d in descs)  # filtro de fornecedor
    row = [r for r in rows if r and "DAS / Simples" in str(r[1])][0]
    assert row[0] == "Fornecedor Alfa"
    assert row[3] == pytest.approx(1330.21)   # valor numérico
    assert row[6] == "Pendente"
    assert row[7] == "DAS / Simples Nacional"  # categoria
    assert row[8] == "Administrativo"          # centro de custo


# ════════════════════════ DESPESAS ════════════════════════

def test_expenses_exports_utf8_and_status_filter(testing_app):
    cid = _cid_of(testing_app)
    cats = _seed_catalog(testing_app, cid)
    _seed_expense(testing_app, cid, cats["prolabore"], cats["cc"],
                  desc="Pró-Labore — competência 07/2026", amount=10000.0,
                  emission=date(2026, 7, 31), due=date(2026, 8, 10),
                  status="pago", paid_date=date(2026, 8, 10))
    _seed_expense(testing_app, cid, cats["das"], cats["cc"],
                  desc="DAS / Simples Nacional — competência 07/2026",
                  amount=1330.21, emission=date(2026, 7, 31), due=date.today() + timedelta(days=10))

    c = _login(testing_app)
    r = c.get("/financial/expenses")
    assert r.status_code == 200
    assert "/financial/expenses/export/pdf" in r.get_data(as_text=True)

    rp = c.get("/financial/expenses/export/pdf")
    assert rp.status_code == 200 and rp.mimetype == "application/pdf"
    assert b"Despesas Gerais" in rp.data
    assert b"10.000" in rp.data and b"1.330,2" in rp.data

    rx = c.get("/financial/expenses/export/xlsx")
    ws, rows = _xlsx_rows(rx)
    assert rows[0][0] == "Despesas Gerais"
    pl = [r for r in rows if r and "Pró-Labore" in str(r[0])][0]
    das = [r for r in rows if r and "DAS / Simples" in str(r[0])][0]
    assert pl[1] == "Pró-Labore" and pl[2] == "Pessoal"
    assert pl[3] == "Administrativo"
    assert pl[6] == 10000.0 and pl[7] == "Paga"
    assert "competência 07/2026" in pl[9]  # observação com acento preservada
    assert das[2] == "Impostos e Tributos" and das[7] == "Pendente"
    assert das[6] == pytest.approx(1330.21)

    rx2 = c.get("/financial/expenses/export/xlsx?status=pago")
    _, rows2 = _xlsx_rows(rx2)
    descs2 = [str(r[0]) for r in rows2 if r and r[0] is not None]
    assert any("Pró-Labore" in d for d in descs2)
    assert not any("DAS / Simples" in d for d in descs2)


# ════════════════════════ LANÇAMENTOS ════════════════════════

def test_lancamentos_exports(testing_app):
    cid = _cid_of(testing_app)
    cats = _seed_catalog(testing_app, cid)
    _seed_expense(testing_app, cid, cats["prolabore"], cats["cc"],
                  desc="Pró-Labore — competência 07/2026", amount=10000.0,
                  emission=date(2026, 7, 31), due=date(2026, 8, 10),
                  status="pago", paid_date=date(2026, 8, 10))
    _seed_expense(testing_app, cid, cats["das"], cats["cc"],
                  desc="DAS / Simples Nacional — competência 07/2026",
                  amount=1330.21, emission=date(2026, 7, 31), due=date.today() + timedelta(days=10))

    c = _login(testing_app)
    r = c.get("/financial/")
    assert r.status_code == 200
    assert "/financial/export/pdf" in r.get_data(as_text=True)

    rp = c.get("/financial/export/pdf?period=all")
    assert rp.status_code == 200 and rp.mimetype == "application/pdf"
    assert b"Lan" in rp.data  # título "Lançamentos Financeiros"

    rx = c.get("/financial/export/xlsx?period=all")
    ws, rows = _xlsx_rows(rx)
    pl = [r for r in rows if r and "Pró-Labore" in str(r[1])][0]
    das = [r for r in rows if r and "DAS / Simples" in str(r[1])][0]
    assert pl[2] == "Despesa" and das[2] == "Despesa"
    assert pl[5] == 10000.0 and das[5] == pytest.approx(1330.21)
    assert pl[7].startswith("expense:") and das[7].startswith("expense:")
    assert pl[8].date() == date(2026, 7, 31)   # competência
    assert pl[9].date() == date(2026, 8, 10)   # pagamento (pago)
    assert das[9] is None or das[9] == "—"  # pendente sem pagamento


# ════════════════════════ CATÁLOGOS (somente XLSX) ════════════════════════

def test_catalogs_exports_and_permissions(testing_app):
    cid = _cid_of(testing_app)
    _seed_catalog(testing_app, cid)

    c = _login(testing_app)
    rx = c.get("/financial/categories/export/xlsx")
    assert rx.status_code == 200
    assert rx.mimetype == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    _, rows = _xlsx_rows(rx)
    cats_ = [str(r[0]).strip() for r in rows if r and r[0] is not None]
    assert "Pessoal" in cats_ and "Pró-Labore" in cats_
    assert "Impostos e Tributos" in cats_ and "DAS / Simples Nacional" in cats_
    row_das = [r for r in rows if r and r[0] and "DAS / Simples" in str(r[0])][0]
    assert row_das[2] == "Impostos e Tributos"  # categoria pai

    rx2 = c.get("/financial/cost-centers/export/xlsx")
    assert rx2.status_code == 200
    _, rows2 = _xlsx_rows(rx2)
    assert any("Administrativo" in str(r[0]) for r in rows2 if r and r[0])

    # Usuário com financial.view mas SEM financial.manage não exporta catálogo
    email = f"viewonly_{uuid.uuid4().hex[:6]}@test.local"
    with testing_app.app_context():
        from app.models.rbac import Role
        viewer_role = Role.query.filter_by(code="VIEWER").first()
        u = User(email=email, name="View Only", company_id=cid,
                 is_active=True, role="operator")
        u.set_password("TestView1!")
        if viewer_role:
            u.roles = [viewer_role]
        db.session.add(u)
        db.session.commit()
    c2 = _login(testing_app, email=email, password="TestView1!")
    assert c2.get("/financial/categories/export/xlsx").status_code == 403
    assert c2.get("/financial/cost-centers/export/xlsx").status_code == 403


# ════════════════════════ SEGURANÇA / EMPRESA / SEM ESCRITA ════════════════════════

def test_exports_require_login(testing_app):
    c = testing_app.test_client()
    for url in ("/financial/cash-flow/export/pdf", "/financial/cash-flow/export/xlsx",
                "/financial/receivables/export/pdf", "/financial/payables/export/xlsx",
                "/financial/expenses/export/pdf", "/financial/export/xlsx",
                "/financial/categories/export/xlsx", "/financial/cost-centers/export/xlsx"):
        r = c.get(url)
        assert r.status_code == 302, url
        assert "/auth/login" in r.headers.get("Location", "")


def test_exports_company_isolation(testing_app):
    cid = _cid_of(testing_app)
    cats = _seed_catalog(testing_app, cid)
    _seed_expense(testing_app, cid, cats["prolabore"], cats["cc"],
                  desc="Despesa Empresa A", amount=1000.0,
                  emission=date(2026, 7, 15), due=date(2026, 8, 15))
    with testing_app.app_context():
        comp_b = Company(name="Empresa B 12E3", slug=f"empresa-b-12e3-{uuid.uuid4().hex[:6]}",
                         document="00.000.000/0007-00")
        db.session.add(comp_b)
        db.session.flush()
        cat_b = FinancialCategory(company_id=comp_b.id, name="Pessoal",
                                  type="expense", active=True)
        cc_b = CostCenter(company_id=comp_b.id, name="Administrativo", active=True)
        db.session.add_all([cat_b, cc_b])
        db.session.flush()
        db.session.add(FinancialRecord(
            company_id=comp_b.id, type="expense", category="outro",
            description="Despesa Empresa B", amount=99999.0, status="pendente",
            emission_date=date(2026, 7, 15), due_date=date(2026, 8, 15),
            financial_category_id=cat_b.id, cost_center_id=cc_b.id,
            reference=f"expense:{uuid.uuid4().hex[:6]}"))
        db.session.commit()

    c = _login(testing_app)
    rx = c.get("/financial/expenses/export/xlsx")
    _, rows = _xlsx_rows(rx)
    descs = [str(r[0]) for r in rows if r and r[0] is not None]
    assert any("Despesa Empresa A" in d for d in descs)
    assert not any("Despesa Empresa B" in d for d in descs)


# ════════════════════════ 12E-A5 — CORREÇÕES (largura e totais) ════════════════════════

def test_xlsx_money_column_width_and_numeric(testing_app):
    """Largura da coluna monetária considera o formato visual (R$ 10.000,00)
    com mínimo de 14 — e o valor continua NUMÉRICO com formato R$."""
    from app.services import report_xlsx
    blob = report_xlsx.build_report_xlsx(
        title="Teste Largura",
        meta_lines=[],
        columns=["Descrição", "Valor"],
        rows=[["R$ 10.000,00", 10000.0],
              ["R$ 1.330,21", 1330.21],
              ["R$ 100.000,00", 100000.0],
              ["R$ 1.000.000,00", 1000000.0]],
        money_cols=(1,))
    wb = load_workbook(BytesIO(blob))
    ws = wb.active
    # cabeçalho na linha 3 (titulo + meta vazia + branco)
    hdr_row = None
    for row in ws.iter_rows():
        if row[0].value == "Descrição":
            hdr_row = row[0].row
            break
    assert hdr_row == 3
    width_b = ws.column_dimensions["B"].width
    assert width_b >= 14, f"largura B = {width_b} (< 14)"
    for r in range(hdr_row + 1, hdr_row + 5):
        cell = ws.cell(row=r, column=2)
        assert isinstance(cell.value, (int, float))
        assert "R$" in (cell.number_format or "")


# ════════════════════════ 12E-A6.1 — SANITIZAÇÃO DE FÓRMULAS ════════════════════════

def test_xlsx_formula_injection_sanitized(testing_app):
    """Textos iniciados por = + - @ viram TEXTO seguro (nunca fórmula)."""
    from app.services import report_xlsx
    dangerous = [
        '=HYPERLINK("https://example.com","Clique")',
        "+1+1",
        "-cmd|' /C calc'!A0",
        "@SUM(1,1)",
    ]
    normal = ["Pró-Labore", "DAS / Simples Nacional", "Impostos e Tributos",
              "Despesas Administrativas", "Anderson Nobre"]
    blob = report_xlsx.build_report_xlsx(
        title="Teste Sanitização", meta_lines=[],
        columns=["Descrição", "Valor"],
        rows=[[t, 1000.0 + i] for i, t in enumerate(dangerous + normal)],
        money_cols=(1,))
    wb = load_workbook(BytesIO(blob))
    ws = wb.active
    # título na linha 1; cabeçalho na 3; dados 4..12
    for i, t in enumerate(dangerous):
        cell = ws.cell(row=4 + i, column=1)
        assert cell.data_type != "f", f"'{t[:20]}' virou fórmula"
        assert cell.value == "'" + t          # prefixo de segurança preservado
        assert cell.value.startswith("'")
    for i, t in enumerate(normal):
        cell = ws.cell(row=4 + len(dangerous) + i, column=1)
        assert cell.value == t, f"texto normal alterado: {cell.value!r}"
    # moeda continua NUMBER com formato R$
    for r in range(4, 4 + len(dangerous) + len(normal)):
        cell = ws.cell(row=r, column=2)
        assert isinstance(cell.value, (int, float))
        assert "R$" in (cell.number_format or "")
    # largura mínima monetária preservada
    assert ws.column_dimensions["B"].width >= 14


def test_xlsx_sanitization_in_real_export(testing_app):
    """Campo real exportável (descrição de despesa) com payload de fórmula
    sai como texto seguro no XLSX de Despesas."""
    cid = _cid_of(testing_app)
    cats = _seed_catalog(testing_app, cid)
    payload = '=HYPERLINK("https://example.com","Clique")'
    _seed_expense(testing_app, cid, cats["prolabore"], cats["cc"],
                  desc=payload, amount=10000.0,
                  emission=date(2026, 7, 31), due=date(2026, 8, 10))

    c = _login(testing_app)
    r = c.get("/financial/expenses/export/xlsx")
    assert r.status_code == 200
    ws, rows = _xlsx_rows(r)
    found = False
    for row in ws.iter_rows():
        if row[0].value == "'" + payload:
            assert row[0].data_type != "f"
            found = True
    assert found, "descrição com payload não encontrada/sanitizada no XLSX"


def test_lancamentos_totals_by_type(testing_app):
    """Lançamentos: totais por tipo (Receitas/Custos/Despesas) + resultado
    líquido — sem total misto ambíguo."""
    cid = _cid_of(testing_app)
    cats = _seed_catalog(testing_app, cid)
    with testing_app.app_context():
        db.session.add_all([
            FinancialRecord(company_id=cid, type="revenue", category="receita_servico",
                            description="Receita teste", amount=5000.0, status="pago",
                            paid_date=date(2026, 8, 5), emission_date=date(2026, 8, 1),
                            reference=f"man:{uuid.uuid4().hex[:6]}"),
            FinancialRecord(company_id=cid, type="cost", category="custo_fornecedor",
                            description="Custo teste", amount=2000.0, status="pago",
                            paid_date=date(2026, 8, 6), emission_date=date(2026, 8, 2),
                            reference=f"man:{uuid.uuid4().hex[:6]}"),
            FinancialRecord(company_id=cid, type="expense", category="outro",
                            description="Despesa teste", amount=1000.0, status="pendente",
                            emission_date=date(2026, 8, 3), due_date=date(2026, 8, 20),
                            financial_category_id=cats["prolabore"],
                            cost_center_id=cats["cc"],
                            reference=f"expense:{uuid.uuid4().hex[:6]}"),
        ])
        db.session.commit()

    c = _login(testing_app)
    rx = c.get("/financial/export/xlsx?period=all")
    assert rx.status_code == 200
    ws, rows = _xlsx_rows(rx)
    labels = [str(r[0]) for r in rows if r and r[0] is not None]
    # sem o antigo total misto
    assert "TOTAIS" not in labels
    tr = _row_of(rows, "TOTAL RECEITAS")
    tc = _row_of(rows, "TOTAL CUSTOS")
    td = _row_of(rows, "TOTAL DESPESAS")
    rl = _row_of(rows, "RESULTADO LÍQUIDO (Receitas − Custos − Despesas)")
    assert tr and tr[5] == 5000.0
    assert tc and tc[5] == 2000.0
    assert td and td[5] == 1000.0
    assert rl and rl[5] == pytest.approx(2000.0)  # 5000 - 2000 - 1000

    rp = c.get("/financial/export/pdf?period=all")
    assert rp.status_code == 200 and rp.mimetype == "application/pdf"
    pdf = rp.data
    # rótulos (texto pode quebrar de linha na coluna estreita)
    assert b"TOTAL" in pdf and b"RECEIT" in pdf
    assert b"CUSTOS" in pdf and b"DESPES" in pdf
    assert b"RESULT" in pdf and b"L" in pdf
    # sem o antigo rótulo de total misto
    assert b"TOTAIS" not in pdf
    # valores por tipo + líquido
    assert b"5.000" in pdf and b"2.000" in pdf and b"1.000" in pdf


def test_exports_do_not_write(testing_app):
    cid = _cid_of(testing_app)
    cats = _seed_catalog(testing_app, cid)
    _seed_expense(testing_app, cid, cats["prolabore"], cats["cc"],
                  desc="Despesa Base", amount=1000.0,
                  emission=date(2026, 7, 15), due=date(2026, 8, 15))

    def _counts():
        with testing_app.app_context():
            return (FinancialRecord.query.count(), Order.query.count(),
                    Client.query.count(), CostCenter.query.count(),
                    FinancialCategory.query.count())

    before = _counts()
    c = _login(testing_app)
    for url in ("/financial/cash-flow", "/financial/cash-flow/export/pdf",
                "/financial/cash-flow/export/xlsx",
                "/financial/receivables", "/financial/receivables/export/pdf",
                "/financial/receivables/export/xlsx",
                "/financial/payables", "/financial/payables/export/pdf",
                "/financial/payables/export/xlsx",
                "/financial/expenses", "/financial/expenses/export/pdf",
                "/financial/expenses/export/xlsx",
                "/financial/", "/financial/export/pdf", "/financial/export/xlsx",
                "/financial/categories", "/financial/categories/export/xlsx",
                "/financial/cost-centers", "/financial/cost-centers/export/xlsx"):
        r = c.get(url)
        assert r.status_code == 200, url
    assert _counts() == before
