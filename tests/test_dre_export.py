"""tests/test_dre_export.py — Exportação da DRE PDF/XLSX (Etapa 12E piloto).

Cobre: tela continua 200; PDF 200 application/pdf com título e valores;
XLSX 200 com mimetype correto, título, valores numéricos monetários e
acentos preservados; filtros da tela respeitados; anônimo bloqueado;
usuário sem financial.view bloqueado; isolamento por empresa; nenhuma
escrita financeira durante a geração.

App próprio com TestingConfig (sqlite :memory:) — não usa o DB dev.
"""
import uuid
from datetime import date, datetime
from io import BytesIO

import pytest
from openpyxl import load_workbook

from app import create_app
from app.extensions import db
from app.models.client import Client
from app.models.company import Company
from app.models.financial import FinancialRecord
from app.models.financial_catalog import CostCenter, FinancialCategory
from app.models.order import Order, OrderPayment
from app.models.user import User

ADMIN_EMAIL = "admin@executivecarsp.com"
ADMIN_PWD = "admin123"

JUL = date(2026, 7, 1)
JUL_END = date(2026, 7, 31)


@pytest.fixture(scope="session")
def testing_app():
    return create_app("testing")


@pytest.fixture(autouse=True)
def clean_tables(testing_app):
    with testing_app.app_context():
        for model in (FinancialRecord, OrderPayment, Order,
                      CostCenter, FinancialCategory, Client):
            model.query.delete()
        db.session.commit()
    yield


def _login(app, email=ADMIN_EMAIL, password=ADMIN_PWD):
    c = app.test_client()
    r = c.post("/auth/login", data={"email": email, "password": password},
               follow_redirects=False)
    assert r.status_code in (200, 302)
    return c


def _cid_of(app, email=ADMIN_EMAIL):
    with app.app_context():
        return User.query.filter_by(email=email).first().company_id


def _seed_catalog(app, cid):
    """Raízes/filhas com acentos (espelham produção) + centro."""
    with app.app_context():
        pessoal = FinancialCategory(company_id=cid, name="Pessoal",
                                    type="expense", active=True)
        db.session.add(pessoal)
        db.session.flush()
        prolabore = FinancialCategory(company_id=cid, name="Pró-Labore",
                                      type="expense", active=True,
                                      parent_id=pessoal.id)
        impostos = FinancialCategory(company_id=cid, name="Impostos e Tributos",
                                     type="expense", active=True)
        db.session.add_all([prolabore, impostos])
        db.session.flush()
        das = FinancialCategory(company_id=cid, name="DAS / Simples Nacional",
                                type="expense", active=True, parent_id=impostos.id)
        cc = CostCenter(company_id=cid, name="Administrativo", active=True)
        db.session.add_all([das, cc])
        db.session.commit()
        return {"prolabore": prolabore.id, "impostos": impostos.id,
                "das": das.id, "cc": cc.id}


def _seed_expense(app, cid, cat_id, cc_id, *, desc, amount, emission,
                  due=date(2026, 8, 10)):
    with app.app_context():
        fr = FinancialRecord(company_id=cid, type="expense", category="outro",
                             description=desc, amount=amount, status="pendente",
                             emission_date=emission, due_date=due,
                             financial_category_id=cat_id, cost_center_id=cc_id,
                             reference=f"expense:{uuid.uuid4().hex[:6]}")
        db.session.add(fr)
        db.session.commit()
        return fr.id


def _seed_order_faturada(app, cid, *, invoiced_at, total):
    with app.app_context():
        client = Client(company_id=cid, name=f"Cliente {uuid.uuid4().hex[:6]}")
        db.session.add(client)
        db.session.flush()
        o = Order(company_id=cid, client_id=client.id,
                  number=f"SO-DREX-{uuid.uuid4().hex[:8]}", status="faturado",
                  client_name=client.name, contact_name="", email="", celular="",
                  language="pt", billing_type="recibo", total_amount=total,
                  payment_method="PIX", emission_date=date.today(),
                  invoiced_at=invoiced_at, created_by=1)
        db.session.add(o)
        db.session.commit()
        return o.id


def _xlsx_rows(resp):
    wb = load_workbook(BytesIO(resp.data))
    ws = wb.active
    return ws, [[c.value for c in row] for row in ws.iter_rows()]


# ─────────────────────────────────────────────────────────────────────────────
# 1–3. Endpoints básicos
# ─────────────────────────────────────────────────────────────────────────────

def test_dre_screen_still_200(testing_app):
    c = _login(testing_app)
    r = c.get("/financial/dre")
    assert r.status_code == 200
    html = r.get_data(as_text=True)
    assert "DRE Gerencial" in html
    # Botões de exportação presentes, com filtros repassados
    assert "/financial/dre/export/pdf" in html
    assert "/financial/dre/export/xlsx" in html
    assert "aria-label=\"Gerar PDF\"" in html
    assert "aria-label=\"Exportar XLSX\"" in html


def test_dre_pdf_ok(testing_app):
    c = _login(testing_app)
    r = c.get("/financial/dre/export/pdf")
    assert r.status_code == 200
    assert r.mimetype == "application/pdf"
    assert r.data.startswith(b"%PDF")
    assert b"DRE Gerencial" in r.data


def test_dre_xlsx_ok(testing_app):
    c = _login(testing_app)
    r = c.get("/financial/dre/export/xlsx")
    assert r.status_code == 200
    assert r.mimetype == ("application/vnd.openxmlformats-officedocument."
                          "spreadsheetml.sheet")
    ws, rows = _xlsx_rows(r)
    assert rows[0][0] == "DRE Gerencial"


# ─────────────────────────────────────────────────────────────────────────────
# 4–5. Valores e acentos (caso real: Pró-Labore 10.000 + DAS 1.330,21)
# ─────────────────────────────────────────────────────────────────────────────

def test_dre_pdf_contains_values(testing_app):
    cid = _cid_of(testing_app)
    cats = _seed_catalog(testing_app, cid)
    _seed_expense(testing_app, cid, cats["prolabore"], cats["cc"],
                  desc="Pró-Labore — competência 07/2026", amount=10000.0,
                  emission=date(2026, 7, 31))
    _seed_expense(testing_app, cid, cats["das"], cats["cc"],
                  desc="DAS / Simples Nacional — competência 07/2026",
                  amount=1330.21, emission=date(2026, 7, 31))
    _seed_order_faturada(testing_app, cid,
                         invoiced_at=datetime(2026, 7, 20, 10, 0), total=5000.0)

    c = _login(testing_app)
    r = c.get("/financial/dre/export/pdf?period=custom&date_from=2026-07-01&date_to=2026-07-31")
    assert r.status_code == 200
    pdf = r.data
    assert b"DRE Gerencial" in pdf
    assert b"01/07/2026" in pdf and b"31/07/2026" in pdf
    assert b"Pessoal" in pdf
    assert b"Impostos" in pdf
    assert b"5.000" in pdf          # receita
    assert b"10.000" in pdf         # pró-labore
    assert b"1.330,2" in pdf        # DAS (kern split: '1.330,2' + '1')
    assert b"RECEITA DE SERVI" in pdf
    assert b"RESULTADO OPERACIONAL" in pdf


def test_dre_xlsx_contains_values(testing_app):
    cid = _cid_of(testing_app)
    cats = _seed_catalog(testing_app, cid)
    _seed_expense(testing_app, cid, cats["prolabore"], cats["cc"],
                  desc="Pró-Labore — competência 07/2026", amount=10000.0,
                  emission=date(2026, 7, 31))
    _seed_expense(testing_app, cid, cats["das"], cats["cc"],
                  desc="DAS / Simples Nacional — competência 07/2026",
                  amount=1330.21, emission=date(2026, 7, 31))
    _seed_order_faturada(testing_app, cid,
                         invoiced_at=datetime(2026, 7, 20, 10, 0), total=5000.0)

    c = _login(testing_app)
    r = c.get("/financial/dre/export/xlsx?period=custom&date_from=2026-07-01&date_to=2026-07-31")
    ws, rows = _xlsx_rows(r)
    assert rows[0][0] == "DRE Gerencial"
    assert any("01/07/2026" in str(row[0]) for row in rows if row[0])  # período
    labels = [str(row[0]) for row in rows if row[0]]
    assert "(−) Pessoal" in labels
    assert "(−) Impostos e Tributos" in labels
    assert "Margem Bruta" in labels
    values = {}
    for row in rows:
        if row[0] in ("Receita de Serviços", "(−) Pessoal",
                      "(−) Impostos e Tributos", "Margem Bruta"):
            values[row[0]] = row[1]
    assert values["Receita de Serviços"] == pytest.approx(5000.0)
    assert values["(−) Pessoal"] == pytest.approx(-10000.0)
    assert values["(−) Impostos e Tributos"] == pytest.approx(-1330.21)
    assert values["Margem Bruta"] == pytest.approx(5000.0)


# ─────────────────────────────────────────────────────────────────────────────
# 6. Filtros da tela respeitados
# ─────────────────────────────────────────────────────────────────────────────

def test_dre_export_respects_period_filter(testing_app):
    cid = _cid_of(testing_app)
    cats = _seed_catalog(testing_app, cid)
    _seed_expense(testing_app, cid, cats["prolabore"], cats["cc"],
                  desc="Despesa Julho", amount=1000.0, emission=date(2026, 7, 15))
    _seed_expense(testing_app, cid, cats["prolabore"], cats["cc"],
                  desc="Despesa Agosto", amount=2000.0, emission=date(2026, 8, 15))

    c = _login(testing_app)
    r = c.get("/financial/dre/export/xlsx?period=custom&date_from=2026-07-01&date_to=2026-07-31")
    _, rows = _xlsx_rows(r)
    all_text = " ".join(str(v) for row in rows for v in row if v is not None)
    assert "Despesa Julho" not in all_text  # detalhe não entra no XLSX piloto
    # Verificação pelos totais: julho deve somar 1000 em Pessoal
    for row in rows:
        if row[0] == "(−) Pessoal":
            assert row[1] == pytest.approx(-1000.0)

    r2 = c.get("/financial/dre/export/xlsx?period=custom&date_from=2026-08-01&date_to=2026-08-31")
    _, rows2 = _xlsx_rows(r2)
    for row in rows2:
        if row[0] == "(−) Pessoal":
            assert row[1] == pytest.approx(-2000.0)


# ─────────────────────────────────────────────────────────────────────────────
# 7–9. Segurança / RBAC / multiempresa
# ─────────────────────────────────────────────────────────────────────────────

def test_dre_export_requires_login(testing_app):
    c = testing_app.test_client()
    for url in ("/financial/dre/export/pdf", "/financial/dre/export/xlsx"):
        r = c.get(url)
        assert r.status_code == 302
        assert "/auth/login" in r.headers.get("Location", "")


def test_dre_export_requires_financial_view(testing_app):
    cid = _cid_of(testing_app)
    email = f"noperm_{uuid.uuid4().hex[:6]}@test.local"
    with testing_app.app_context():
        u = User(email=email, name="Sem Perm", company_id=cid,
                 is_active=True, role="operator")
        u.set_password("TestNoPerm1!")
        db.session.add(u)
        db.session.commit()
    c = _login(testing_app, email=email, password="TestNoPerm1!")
    for url in ("/financial/dre/export/pdf", "/financial/dre/export/xlsx"):
        r = c.get(url)
        assert r.status_code == 403


def test_dre_export_company_isolation(testing_app):
    cid = _cid_of(testing_app)
    cats = _seed_catalog(testing_app, cid)
    _seed_expense(testing_app, cid, cats["prolabore"], cats["cc"],
                  desc="Despesa Empresa A", amount=1000.0, emission=date(2026, 7, 15))
    with testing_app.app_context():
        comp_b = Company(name="Empresa B 12E", slug=f"empresa-b-12e-{uuid.uuid4().hex[:6]}",
                         document="00.000.000/0006-00")
        db.session.add(comp_b)
        db.session.flush()
        cat_b = FinancialCategory(company_id=comp_b.id, name="Pessoal",
                                  type="expense", active=True)
        cc_b = CostCenter(company_id=comp_b.id, name="Administrativo", active=True)
        db.session.add_all([cat_b, cc_b])
        db.session.flush()
        db.session.add(FinancialRecord(
            company_id=comp_b.id, type="expense", category="outro",
            description="Despesa Empresa B", amount=99999.0, status="pendente",
            emission_date=date(2026, 7, 15), due_date=date(2026, 8, 10),
            financial_category_id=cat_b.id, cost_center_id=cc_b.id,
            reference=f"expense:{uuid.uuid4().hex[:6]}"))
        db.session.commit()

    c = _login(testing_app)
    r = c.get("/financial/dre/export/xlsx?period=custom&date_from=2026-07-01&date_to=2026-07-31")
    _, rows = _xlsx_rows(r)
    for row in rows:
        if row[0] == "(−) Pessoal":
            assert row[1] == pytest.approx(-1000.0)  # só empresa A


# ─────────────────────────────────────────────────────────────────────────────
# 10–12. Acentos, números monetários e ausência de escrita
# ─────────────────────────────────────────────────────────────────────────────

def test_xlsx_preserves_special_chars_and_numeric_money(testing_app):
    cid = _cid_of(testing_app)
    cats = _seed_catalog(testing_app, cid)
    _seed_expense(testing_app, cid, cats["prolabore"], cats["cc"],
                  desc="Pró-Labore — competência 07/2026", amount=10000.0,
                  emission=date(2026, 7, 31))
    c = _login(testing_app)
    r = c.get("/financial/dre/export/xlsx?period=custom&date_from=2026-07-01&date_to=2026-07-31")
    ws, rows = _xlsx_rows(r)

    # Meta com acentos preservados
    meta_text = " ".join(str(row[0]) for row in rows[:4] if row[0])
    assert "Período (competência)" in meta_text

    # Grupo com valor numérico e formato monetário
    found = False
    for row in ws.iter_rows():
        if row[0].value == "(−) Pessoal":
            vcell = row[1]
            assert isinstance(vcell.value, (int, float))
            assert vcell.value == pytest.approx(-10000.0)
            assert "R$" in (vcell.number_format or "")
            found = True
    assert found

    # Builder preserva acentos/em-dash em texto livre (nível unitário)
    from app.services import report_xlsx
    blob = report_xlsx.build_report_xlsx(
        title="Teste", meta_lines=["Pró-Labore — competência 07/2026"],
        columns=["Descrição"], rows=[["Pró-Labore — competência 07/2026"]])
    wb2 = load_workbook(BytesIO(blob))
    ws2 = wb2.active
    assert ws2["A2"].value == "Pró-Labore — competência 07/2026"


def test_export_does_not_write_financial_data(testing_app):
    cid = _cid_of(testing_app)
    cats = _seed_catalog(testing_app, cid)
    _seed_expense(testing_app, cid, cats["prolabore"], cats["cc"],
                  desc="Despesa Base", amount=1000.0, emission=date(2026, 7, 15))

    def _counts():
        with testing_app.app_context():
            return (FinancialRecord.query.count(),
                    Order.query.count(),
                    Client.query.count())

    before = _counts()
    c = _login(testing_app)
    for url in ("/financial/dre",
                "/financial/dre/export/pdf",
                "/financial/dre/export/xlsx",
                "/financial/dre/export/pdf?period=custom&date_from=2026-07-01&date_to=2026-07-31",
                "/financial/dre/export/xlsx?period=custom&date_from=2026-07-01&date_to=2026-07-31"):
        r = c.get(url)
        assert r.status_code in (200, 302)
    after = _counts()
    assert before == after
