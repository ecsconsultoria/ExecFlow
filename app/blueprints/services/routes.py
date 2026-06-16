from flask import render_template, jsonify, request, redirect, url_for, flash, send_file, abort
from flask_login import login_required, current_user
from . import services_bp
from ...models.service import Service, State, ServicePricing
from ...models.vehicle import VehicleCategory
from ...extensions import db
from ...utils.decorators import require_permission
from ...utils.audit import log_activity

DRIVER_TYPES = ["Monolíngue", "Bilíngue", ""]


def _get_pricing_or_404(pid: int) -> ServicePricing:
    """Carrega ServicePricing garantindo que pertence ao tenant atual.

    Aceita Service.company_id == current_user.company_id OU None (catálogo
    global compartilhado). Bloqueia acesso a pricing de outras empresas.
    """
    p = (ServicePricing.query
         .join(Service, ServicePricing.service_id == Service.id)
         .filter(ServicePricing.id == pid)
         .filter((Service.company_id == current_user.company_id) |
                 (Service.company_id.is_(None)))
         .first())
    if p is None:
        abort(404)
    return p


def _build_rows(f_service="", f_vehicle="", f_driver="", f_state=""):
    q = (ServicePricing.query
         .join(Service, ServicePricing.service_id == Service.id)
         .join(VehicleCategory, ServicePricing.category_id == VehicleCategory.id)
         .filter(Service.is_active == True, ServicePricing.is_active == True)
         .filter((Service.company_id == current_user.company_id) |
                 (Service.company_id.is_(None))))
    if f_service:
        q = q.filter(Service.name.ilike(f"%{f_service}%"))
    if f_vehicle:
        q = q.filter(VehicleCategory.name.ilike(f"%{f_vehicle}%"))
    if f_driver:
        q = q.filter(ServicePricing.driver_type == f_driver)
    if f_state:
        st = State.query.filter_by(code=f_state).first()
        if st:
            q = q.filter(Service.state_id == st.id)
    return q.order_by(Service.name, VehicleCategory.name, ServicePricing.driver_type).all()


@services_bp.route("/", methods=["GET"])
@login_required
@require_permission("catalog.view")
def index():
    f_service = request.args.get("f_service", "")
    f_vehicle = request.args.get("f_vehicle", "")
    f_driver  = request.args.get("f_driver",  "")
    f_state   = request.args.get("f_state",   "")

    pricing_rows    = _build_rows(f_service, f_vehicle, f_driver, f_state)
    states          = State.query.order_by(State.code).all()
    categories      = VehicleCategory.query.filter_by(is_active=True).order_by(VehicleCategory.name).all()
    service_objects = (Service.query
                       .filter((Service.company_id == current_user.company_id) |
                               (Service.company_id.is_(None)))
                       .filter_by(is_active=True)
                       .order_by(Service.name)
                       .all())
    return render_template("services/index.html",
                           pricing_rows=pricing_rows,
                           states=states, categories=categories,
                           driver_types=DRIVER_TYPES,
                           f_service=f_service, f_vehicle=f_vehicle,
                           f_driver=f_driver, f_state=f_state,
                           service_objects=service_objects)


@services_bp.route("/add", methods=["POST"])
@login_required
@require_permission("catalog.manage")
def add():
    service_name = (request.form.get("service_name") or "").strip()
    category_id  = request.form.get("category_id", type=int)
    driver_type  = (request.form.get("driver_type") or "").strip()
    state_code   = (request.form.get("state_code") or "SP").strip()
    price_cost   = float(request.form.get("price_cost") or 0)
    price_base   = float(request.form.get("price_base") or 0)

    if not service_name or not category_id:
        flash("Tipo de serviço e Veículo são obrigatórios.", "danger")
        return redirect(url_for("services.index"))

    state = State.query.filter_by(code=state_code).first()
    if not state:
        flash(f"Estado '{state_code}' não encontrado.", "danger")
        return redirect(url_for("services.index"))

    svc = (Service.query
           .filter_by(name=service_name, state_id=state.id, is_active=True)
           .filter((Service.company_id == current_user.company_id) | (Service.company_id.is_(None)))
           .first())
    if not svc:
        svc = Service(company_id=current_user.company_id, state_id=state.id,
                      name=service_name, is_active=True)
        db.session.add(svc)
        db.session.flush()

    if ServicePricing.query.filter_by(service_id=svc.id, category_id=category_id,
                                      driver_type=driver_type).first():
        flash("Já existe um registro para esse serviço/veículo/motorista.", "warning")
        return redirect(url_for("services.index"))

    db.session.add(ServicePricing(service_id=svc.id, category_id=category_id,
                                  driver_type=driver_type, price_cost=price_cost,
                                  price_base=price_base, is_active=True))
    db.session.commit()
    flash("Serviço adicionado com sucesso.", "success")
    return redirect(url_for("services.index"))


@services_bp.route("/edit/<int:pid>", methods=["POST"])
@login_required
@require_permission("catalog.manage")
def edit(pid):
    p = _get_pricing_or_404(pid)
    p.price_cost      = float(request.form.get("price_cost")      or 0)
    p.price_base      = float(request.form.get("price_base")      or 0)
    p.price_nf        = float(request.form.get("price_nf")        or 0)
    p.price_cartao    = float(request.form.get("price_cartao")    or 0)
    p.price_nf_cartao = float(request.form.get("price_nf_cartao") or 0)
    db.session.commit()
    flash("Preços atualizados.", "success")
    return redirect(url_for("services.index"))


@services_bp.route("/flags/<int:sid>", methods=["POST"])
@login_required
@require_permission("catalog.manage")
def edit_flags(sid):
    """Toggle individual operational flags on a Service."""
    svc = (Service.query
           .filter_by(id=sid)
           .filter((Service.company_id == current_user.company_id) |
                   (Service.company_id.is_(None)))
           .first_or_404())
    flag = request.json.get("flag") if request.is_json else request.form.get("flag")
    val  = request.json.get("value") if request.is_json else (request.form.get("value") == "true")
    allowed = {"is_operational", "requires_route", "requires_passenger",
               "requires_vehicle", "requires_dispatch", "requires_schedule"}
    if flag not in allowed:
        return ({"error": "invalid flag"}, 400) if request.is_json else ("invalid", 400)
    setattr(svc, flag, bool(val))
    db.session.commit()
    if request.is_json:
        return {"ok": True, "flag": flag, "value": bool(val)}
    flash(f"Flag '{flag}' atualizada.", "success")
    return redirect(url_for("services.index"))


@services_bp.route("/delete/<int:pid>", methods=["POST"])
@login_required
@require_permission("catalog.manage")
def delete(pid):
    p = _get_pricing_or_404(pid)
    p.is_active = False
    svc = p.service
    if svc and not svc.pricing.filter_by(is_active=True).filter(ServicePricing.id != pid).count():
        svc.is_active = False
    log_activity("service_pricing", pid, current_user.company_id, f"Preço de serviço {svc.name if svc else pid} desativado", current_user.id)
    db.session.commit()
    flash("Serviço removido.", "success")
    return redirect(url_for("services.index"))


@services_bp.route("/delete-bulk", methods=["POST"])
@login_required
@require_permission("catalog.manage")
def delete_bulk():
    ids = request.form.getlist("ids")
    skipped = 0
    for raw in ids:
        try:
            pid_int = int(raw)
        except (ValueError, TypeError):
            continue
        # Tenant guard: pula silenciosamente IDs de outras empresas
        p = (ServicePricing.query
             .join(Service, ServicePricing.service_id == Service.id)
             .filter(ServicePricing.id == pid_int)
             .filter((Service.company_id == current_user.company_id) |
                     (Service.company_id.is_(None)))
             .first())
        if not p:
            skipped += 1
            continue
        try:
            p.is_active = False
            svc = p.service
            if svc and not svc.pricing.filter_by(is_active=True).filter(ServicePricing.id != p.id).count():
                svc.is_active = False
        except (ValueError, TypeError):
            pass
    log_activity("service_pricing", 0, current_user.company_id, f"Remoção em lote de {len(ids)} preços de serviço", current_user.id)
    db.session.commit()
    flash(f"{len(ids)} registro(s) removido(s).", "success")
    return redirect(url_for("services.index"))


@services_bp.route("/import-excel", methods=["POST"])
@login_required
@require_permission("catalog.manage")
def import_excel():
    try:
        import openpyxl
    except ImportError:
        flash("openpyxl não instalado. Execute: pip install openpyxl", "danger")
        return redirect(url_for("services.index"))

    f = request.files.get("file")
    if not f or not f.filename:
        flash("Nenhum arquivo enviado.", "danger")
        return redirect(url_for("services.index"))

    try:
        wb = openpyxl.load_workbook(f, data_only=True)
        ws = wb.active
        added, errors = 0, []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or row[0] is None:
                continue
            try:
                service_name = str(row[0]).strip()
                vehicle_name = str(row[1]).strip() if row[1] else ""
                driver_type  = str(row[2]).strip() if row[2] else ""
                state_code   = str(row[3]).strip() if row[3] else "SP"
                price_cost   = float(row[4] or 0)
                price_base   = float(row[5] or 0)

                state = State.query.filter_by(code=state_code).first()
                if not state:
                    errors.append(f"Estado '{state_code}' não encontrado"); continue
                cat = VehicleCategory.query.filter(VehicleCategory.name.ilike(vehicle_name)).first()
                if not cat:
                    errors.append(f"Veículo '{vehicle_name}' não encontrado"); continue

                svc = Service.query.filter_by(name=service_name, state_id=state.id, is_active=True).first()
                if not svc:
                    svc = Service(company_id=current_user.company_id, state_id=state.id,
                                  name=service_name, is_active=True)
                    db.session.add(svc)
                    db.session.flush()

                existing = ServicePricing.query.filter_by(
                    service_id=svc.id, category_id=cat.id, driver_type=driver_type).first()
                if existing:
                    existing.price_cost = price_cost
                    existing.price_base = price_base
                else:
                    db.session.add(ServicePricing(service_id=svc.id, category_id=cat.id,
                                                  driver_type=driver_type, price_cost=price_cost,
                                                  price_base=price_base, is_active=True))
                added += 1
            except Exception as e:
                errors.append(str(e))

        db.session.commit()
        msg = f"{added} registro(s) importado(s)."
        if errors:
            msg += f" Erros: {'; '.join(errors[:5])}"
        flash(msg, "success" if not errors else "warning")
    except Exception as e:
        flash(f"Erro ao importar: {e}", "danger")

    return redirect(url_for("services.index"))


@services_bp.route("/export-excel")
@login_required
@require_permission("catalog.view")
def export_excel():
    """Export all active service pricings to xlsx."""
    import io, openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    rows = _build_rows()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Serviços"

    headers = ["Serviço", "Veículo", "Motorista", "Estado", "Custo (R$)", "Preço Base (R$)",
                "Preço NF (R$)", "Preço Cartão (R$)", "Preço NF+Cartão (R$)"]
    ws.append(headers)
    hdr_fill = PatternFill("solid", fgColor="0b0b0b")
    hdr_font = Font(bold=True, color="FFFFFF")
    for cell in ws[1]:
        cell.fill = hdr_fill
        cell.font = hdr_font
        cell.alignment = Alignment(horizontal="center")

    for p in rows:
        ws.append([
            p.service.name,
            p.category.name if p.category else "",
            p.driver_type or "",
            (p.service.state.code if p.service.state else ""),
            p.price_cost or 0,
            p.price_base or 0,
            p.price_nf or 0,
            p.price_cartao or 0,
            p.price_nf_cartao or 0,
        ])

    for col in ws.columns:
        max_len = max((len(str(cell.value or "")) for cell in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    from datetime import datetime
    fname = f"Servicos_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return send_file(buf, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                     as_attachment=True, download_name=fname)


@services_bp.route("/pricing", methods=["GET"])
@login_required
@require_permission("catalog.view")
def pricing():
    """JSON endpoint: pricing for quote builder and PO item cost lookup."""
    svc_id      = request.args.get("service_id", type=int)
    cat_id      = request.args.get("category_id", type=int)
    billing     = request.args.get("billing_type", "recibo")
    driver_type = request.args.get("driver_type", "").strip()
    if not svc_id or not cat_id:
        return jsonify({"price": 0, "price_cost": 0})
    # Busca exata por service + category + driver_type
    q = ServicePricing.query.filter_by(service_id=svc_id, category_id=cat_id, is_active=True)
    if driver_type:
        p = q.filter_by(driver_type=driver_type).first()
    else:
        p = q.first()
    if not p:
        return jsonify({"price": 0, "price_cost": 0})
    return jsonify({"price": p.effective_price(billing), "price_cost": p.price_cost or 0,
                    "price_base": p.price_base, "price_nf": p.price_nf,
                    "price_cartao": p.price_cartao, "price_nf_cartao": p.price_nf_cartao,
                    "km_extra_rate": p.category.km_extra_rate if p.category else 0})
