"""report_xlsx.py — Relatórios financeiros em XLSX (Etapa 12E).

Wrapper openpyxl com o padrão de exportação do ExecFlow:

  * título (bold) + período + filtros utilizados + data/hora + usuário;
  * cabeçalho destacado (fundo escuro, texto branco) com bordas;
  * valores monetários como NÚMEROS com formato '"R$" #,##0.00' (negativos
    em vermelho) — nunca strings;
  * datas como date com formato dd/mm/aaaa;
  * totais em linha destacada (bold + borda superior dupla);
  * autofiltro no cabeçalho e congelamento (freeze_panes);
  * larguras de coluna calculadas pelo conteúdo.

SOMENTE LEITURA: apenas formata dados já calculados.
"""
from __future__ import annotations

import io
from datetime import date, datetime

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

_C_SLATE = "334155"     # cabeçalho
_C_ZEBRA = "F1F5F9"     # zebra (aplicado em células de seção, se desejado)
_C_TOTAL = "E2E8F0"     # linha de totais

_FMT_BRL = '"R$" #,##0.00;[Red]-"R$" #,##0.00'
_FMT_DATE = "dd/mm/yyyy"

_thin = Side(style="thin", color="CBD5E1")
_double = Side(style="double", color="334155")
_border_all = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)


def _header_fill() -> PatternFill:
    return PatternFill("solid", fgColor=_C_SLATE)


def _auto_widths(columns: list[str], rows: list[list], totals, *,
                 money_cols: tuple[int, ...] = (),
                 date_cols: tuple[int, ...] = ()) -> dict:
    """Largura por coluna a partir do maior conteúdo (limitado a 60).

    Colunas monetárias: largura calculada pelo valor FORMATADO
    ("R$ 10.000,00" ≈ 12–13 chars), com mínimo de 14 — nunca pelo número
    bruto (o formato exibido pelo Excel é mais largo que o valor).
    Colunas de data: mínimo de 12 ("dd/mm/aaaa").
    """
    widths: dict[int, int] = {}
    texts = [columns] + rows
    totals_rows = _norm_totals(totals)
    if totals_rows:
        texts = texts + totals_rows
    for c in range(len(columns)):
        best = 10
        for row in texts:
            if c < len(row):
                v = row[c]
                if v is None:
                    continue
                if c in money_cols:
                    try:
                        s = f"R$ {abs(round(float(v), 2)):,.2f}"
                    except (TypeError, ValueError):
                        s = str(v)
                else:
                    s = str(v)
                best = max(best, min(len(s) + 2, 60))
        if c in money_cols:
            best = max(best, 14)
        elif c in date_cols:
            best = max(best, 12)
        widths[c] = best
    return widths


def _norm_totals(totals):
    """totals pode ser uma linha única ou uma lista de linhas de totais."""
    if totals is None:
        return []
    if totals and isinstance(totals[0], (list, tuple)):
        return list(totals)
    return [list(totals)]


def build_report_xlsx(*, title: str, meta_lines: list[str], columns: list[str],
                      rows: list[list], totals=None,
                      money_cols: tuple[int, ...] = (),
                      date_cols: tuple[int, ...] = (),
                      generated_by: str = "",
                      now: datetime | None = None) -> bytes:
    """Gera o XLSX e retorna os bytes.

    rows: células str | float | date | None. Células de money_cols são
    gravadas como número com formato monetário; date_cols como data.
    totals: uma linha (lista) ou várias linhas de totais (lista de listas) —
    todas destacadas (bold + fundo + borda superior dupla).
    """
    now = now or datetime.now()
    wb = Workbook()
    ws = wb.active
    ws.title = "Relatório"

    ncols = len(columns)
    meta_start = 1
    ws.cell(row=meta_start, column=1, value=title).font = Font(bold=True, size=13)
    row_i = meta_start + 1
    for line in meta_lines:
        c = ws.cell(row=row_i, column=1, value=line)
        c.font = Font(size=9, color="64748B")
        row_i += 1
    row_i += 1  # linha em branco

    header_row = row_i
    for idx, name in enumerate(columns, start=1):
        c = ws.cell(row=header_row, column=idx, value=name)
        c.font = Font(bold=True, color="FFFFFF", size=10)
        c.fill = _header_fill()
        c.border = _border_all
        c.alignment = Alignment(horizontal="right" if (idx - 1) in money_cols else "left",
                                vertical="center")
    row_i += 1

    def _write_cell(r, col, value, *, bold=False, total_row=False):
        cell = ws.cell(row=r, column=col)
        idx = col - 1
        if idx in money_cols and value is not None:
            try:
                num = round(float(value), 2)
            except (TypeError, ValueError):
                num = 0.0
            cell.value = num
            cell.number_format = _FMT_BRL
            cell.alignment = Alignment(horizontal="right")
        elif idx in date_cols and value is not None:
            cell.value = value if isinstance(value, (date, datetime)) else value
            cell.number_format = _FMT_DATE
            cell.alignment = Alignment(horizontal="center")
        else:
            # Sanitização anti formula-injection (12E-A6.1): texto iniciado
            # por = + - @ é prefixado com ' (o Excel exibe o texto literal e
            # NÃO o interpreta como fórmula). Apenas células TEXTUAIS —
            # moeda e datas acima permanecem numéricas/datas.
            text = str(value) if value is not None else "—"
            if text.lstrip().startswith(("=", "+", "-", "@")):
                text = "'" + text
            cell.value = text
        cell.border = _border_all
        if bold:
            cell.font = Font(bold=True)
        if total_row:
            cell.fill = PatternFill("solid", fgColor=_C_TOTAL)
        return cell

    for r, row in enumerate(rows):
        for col in range(1, ncols + 1):
            val = row[col - 1] if col - 1 < len(row) else None
            _write_cell(row_i, col, val)
        row_i += 1

    for total_row in _norm_totals(totals):
        for col in range(1, ncols + 1):
            val = total_row[col - 1] if col - 1 < len(total_row) else None
            cell = _write_cell(row_i, col, val, bold=True, total_row=True)
            if col == 1:
                cell.border = Border(left=_thin, right=_thin, top=_double, bottom=_thin)
            else:
                cell.border = Border(left=_thin, right=_thin, top=_double, bottom=_thin)
        row_i += 1

    # Autofiltro + congelamento do cabeçalho
    last_data_row = max(header_row, row_i - 1)
    ws.auto_filter.ref = f"A{header_row}:{get_column_letter(ncols)}{last_data_row}"
    ws.freeze_panes = ws.cell(row=header_row + 1, column=1)

    for col, w in _auto_widths(columns, rows, totals,
                               money_cols=money_cols, date_cols=date_cols).items():
        ws.column_dimensions[get_column_letter(col + 1)].width = w

    # Metadado de geração (última linha, discreto)
    gen = f"Gerado em {now.strftime('%d/%m/%Y %H:%M')}"
    if generated_by:
        gen += f" por {generated_by}"
    gcell = ws.cell(row=row_i + 1, column=1, value=gen)
    gcell.font = Font(size=8, italic=True, color="94A3B8")

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
